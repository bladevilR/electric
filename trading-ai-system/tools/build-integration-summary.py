from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
APP_ROOT = Path(__file__).resolve().parents[1]
OUTPUT = APP_ROOT / "data" / "integration-summary.json"
STANDARD_SUMMARY = ROOT / "jspec-capture" / "output" / "session-20260507-101645" / "standard" / "dataset-summary.json"
PARTICIPANT_RESPONSE = ROOT / "jspec-capture" / "output" / "session-20260507-101645" / "responses" / "007-www-jspec-com-cn-px-spotgoods-province-mosenergybidinfouser-getparticipants.json"


def number(value):
    return value if isinstance(value, (int, float)) else None


def round_number(value, digits=4):
    if value is None:
        return None
    return round(float(value), digits)


def compact_name(path: Path) -> str:
    return path.name


def parse_ledger(root: Path):
    files = sorted(root.glob("*交易电量*结算一览表.xlsx"))
    if not files:
        return {"sourceFile": None, "months": []}

    file_path = files[0]
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb["2026年"]
    months = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        month = row[0]
        if not isinstance(month, str) or not month.endswith("月"):
            continue
        actual_energy = number(row[31]) if len(row) > 31 else None
        if actual_energy is None:
            continue
        months.append(
            {
                "month": month,
                "actualSettlementEnergyWanKwh": round_number(actual_energy),
                "settlementPriceYuanPerKwh": round_number(row[30] if len(row) > 30 else None, 6),
                "gridProxyPriceYuanPerKwh": round_number(row[8] if len(row) > 8 else None, 6),
                "spotShare": round_number(row[32] if len(row) > 32 else None, 6),
                "savingVsMarketUserWanYuan": round_number(row[38] if len(row) > 38 else None, 4),
                "savingVsMarketTotalWanYuan": round_number(row[40] if len(row) > 40 else None, 4),
                "savingVsGridWanYuan": round_number(row[42] if len(row) > 42 else None, 4),
            }
        )
    return {"sourceFile": compact_name(file_path), "months": months}


def parse_month_from_title(value) -> str | None:
    text = str(value or "")
    match = re.search(r"(\d{4})年(\d{1,2})月", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}"
    return None


def sheet_rows(ws):
    return list(ws.iter_rows(values_only=True))


def header_index(rows, label: str, default: int | None = None) -> int | None:
    header_a = rows[4] if len(rows) > 4 else []
    header_b = rows[5] if len(rows) > 5 else []
    candidates = []
    width = max(len(header_a), len(header_b))
    for index in range(width):
        text = f"{header_a[index] if index < len(header_a) and header_a[index] else ''}{header_b[index] if index < len(header_b) and header_b[index] else ''}"
        if label in text:
            candidates.append(index)
    return candidates[-1] if candidates else default


def parse_check_file(file_path: Path):
    wb = load_workbook(file_path, data_only=True, read_only=True)
    month_sheet = next((name for name in wb.sheetnames if "月" in name), wb.sheetnames[0])
    rows = sheet_rows(wb[month_sheet])
    month = parse_month_from_title(rows[0][0] if rows and rows[0] else None) or month_sheet
    total_fee_index = header_index(rows, "交易电费", default=14)
    if total_fee_index is None:
        total_fee_index = header_index(rows, "总电费", default=11)
    saving_index = header_index(rows, "交易节约费用", default=15)
    total_usage = 0.0
    total_fee = 0.0
    total_saving = 0.0
    point_rows = 0
    for row in rows:
        if not row or not re.match(r"^\d{2}:\d{2}$", str(row[0] or "")):
            continue
        point_rows += 1
        total_usage += float(number(row[1]) or 0)
        if total_fee_index is not None and total_fee_index < len(row):
            total_fee += float(number(row[total_fee_index]) or 0)
        if saving_index is not None and saving_index < len(row):
            total_saving += float(number(row[saving_index]) or 0)
    daily_sheets = [
        name
        for name in wb.sheetnames
        if name.isdigit() and sheet_rows(wb[name]) and re.search(r"\d{4}年\d{1,2}月\d{1,2}日", str(sheet_rows(wb[name])[0][0] or ""))
    ]
    return {
        "fileName": compact_name(file_path),
        "month": month,
        "dailySheets": len(daily_sheets),
        "monthlyPointRows": point_rows,
        "totalUsageMwh": round_number(total_usage),
        "totalTradeFeeYuan": round_number(total_fee, 2),
        "totalSavingYuan": round_number(total_saving, 2),
    }


def parse_settlement_checks(root: Path):
    files = sorted(root.glob("*现货核对单*.xlsx"))
    return {"files": [parse_check_file(path) for path in files]}


def parse_participants():
    if not PARTICIPANT_RESPONSE.exists():
        return []
    payload = json.loads(PARTICIPANT_RESPONSE.read_text(encoding="utf-8"))
    return payload.get("bodyJson", {}).get("data", []) or []


def main():
    standard = json.loads(STANDARD_SUMMARY.read_text(encoding="utf-8"))
    summary = {
        "generatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "ledger": parse_ledger(ROOT),
        "settlementChecks": parse_settlement_checks(ROOT),
        "participants": parse_participants(),
        "standard": {
            "p0SourceCoverage": {
                "present": len([source for source in standard.get("sources", {}).values() if source]),
                "total": len(standard.get("sources", {})),
            },
            "gaps": standard.get("gaps", []),
            "fieldCompleteness": standard.get("fieldCompleteness", {}),
        },
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(OUTPUT)


if __name__ == "__main__":
    main()
